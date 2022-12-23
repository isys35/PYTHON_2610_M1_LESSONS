import openpyxl
import pandas
import csv

wb = openpyxl.load_workbook("test_excel.xlsx")
sheet = wb['168 x 7I3A 23.12']
# print(sheet.title)
cell = sheet["D4"]

# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

cell_2 = sheet.cell(4, 5)

# print(cell_2.value)
# for i in range(2, 20):
#     print(i, sheet.cell(4, i).value)

MAX_ROW_LENGTH = 5
result_data = []
row_data = []

target = 0

for row in range(1, sheet.max_row):
    for column in range(1, sheet.max_column):
        if isinstance(sheet.cell(row, column).value, int):
            target += 1
            if target == 6:
                target = 0
                row_data = []
                result_data.append(row_data)
            else:
                row_data.append(sheet.cell(row, column).value)


# with open("result_excel.csv", "w") as csv_file:
#     csv.writer(csv_file, delimiter=";").writerows(result_data)


# data_frame = pandas.DataFrame(sheet.values)
# print(data_frame)









